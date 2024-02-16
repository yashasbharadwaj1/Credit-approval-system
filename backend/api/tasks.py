from celery import shared_task
from api.models import Customer, Loan
import os
import openpyxl


@shared_task()
def populate_customer_data_from_excel():

    file_name = "customer_data.xlsx"
    file_path = os.path.join(os.getcwd(), "data", file_name)

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if all(
            cell is not None for cell in row
        ):  # Check if all cells in the row are not None
            customer = Customer(
                customer_id=row[0],
                first_name=row[1],
                last_name=row[2],
                age=row[3],
                phone_number=row[4],
                monthly_salary=row[5],
                approved_limit=row[6],
            )
            customer.save()

    return f"Customer data loaded successfully"


@shared_task()
def populate_loan_data_from_excel():

    file_name = "loan_data.xlsx"
    file_path = os.path.join(os.getcwd(), "data", file_name)

    workbook = openpyxl.load_workbook(file_path)
    worksheet = workbook.active

    for row in worksheet.iter_rows(min_row=2, values_only=True):
        if all(
            cell is not None for cell in row
        ):  # Check if all cells in the row are not None
            loan = Loan(
                customer_id=row[0],
                loan_id=row[1],
                loan_amount=row[2],
                tenure=row[3],
                interest_rate=row[4],
                monthly_payment=row[5],
                emis_paid_on_time=row[6],
                date_of_approval=row[7],
                end_date=row[8],
            )
            loan.save()

    return f"Loan data Loaded successfully"
